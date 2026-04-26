from __future__ import annotations

import json
import re
import subprocess
import sys
import time
from pathlib import Path

import requests

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(r"D:\Clawdbot_Docker_20260125")
CONSUME_DIR = ROOT / "clawstack_v2" / "data" / "paperless" / "consume"
SOURCE_DIR = next(p for p in CONSUME_DIR.iterdir() if p.name.startswith("IATF") and 25104 in [ord(ch) for ch in p.name])
TEMPLATE_2024 = SOURCE_DIR / "プロセスの監視・測定記録_提出.xlsx"
PDF_YEAR_DIRS = {
    int(p.name): p
    for p in SOURCE_DIR.iterdir()
    if p.is_dir() and p.name.isdigit() and int(p.name) >= 2025
}
OUTPUT_JSON = ROOT / "iatf_system" / "db" / "process_monitoring_measurement.json"
MAX_TEMPLATE_COL = 24

DOCLING_URL = "http://localhost:8087/v1/convert/file"
MONTH_PATTERN = re.compile(r"(20\d{2})年(\d{1,2})月度")
DATE_PATTERN = re.compile(r"作成日\s*[:：]\s*(\d{4}年\d{1,2}月\d{1,2}日)")
PROCESS_KEYWORDS = ["プロセス"]
SECTION_STOP = re.compile(r"^[ⅢⅣⅤ]")
TARGET_RE = re.compile(r"(以上|以下|達成|件[/／]年|件以上|以内|%[/／]年|％[/／]年|100%|１００％|0件|1件|2件|4件)")
ACTUAL_RE = re.compile(r"^(当月|内部監査|累計|校正)")


# ────────────────────────────────────────────────
# Excel 2024 テンプレート抽出（既存ロジック）
# ────────────────────────────────────────────────

def normalize_cell(value):
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.4f}".rstrip("0").rstrip(".")
    return str(value).replace("\r\n", "\n").replace("\r", "\n")


def extract_2024():
    if not TEMPLATE_2024.exists():
        raise FileNotFoundError(f"Template not found: {TEMPLATE_2024}")

    wb = load_workbook(TEMPLATE_2024, data_only=True)
    ws = wb[wb.sheetnames[0]]
    max_col = min(ws.max_column, MAX_TEMPLATE_COL)

    covered = set()
    merged_ranges = []
    for merged in ws.merged_cells.ranges:
        min_col, min_row, merged_max_col, max_row = merged.bounds
        if min_col > MAX_TEMPLATE_COL:
            continue
        merged_max_col = min(merged_max_col, MAX_TEMPLATE_COL)
        merged_ranges.append({
            "start_row": min_row, "start_col": min_col,
            "end_row": max_row, "end_col": merged_max_col,
            "rowspan": max_row - min_row + 1,
            "colspan": merged_max_col - min_col + 1,
        })
        for row in range(min_row, max_row + 1):
            for col in range(min_col, merged_max_col + 1):
                if row == min_row and col == min_col:
                    continue
                covered.add((row, col))

    merged_lookup = {(m["start_row"], m["start_col"]): m for m in merged_ranges}
    rows = []
    for row_idx in range(1, ws.max_row + 1):
        row_cells = []
        for col_idx in range(1, max_col + 1):
            if (row_idx, col_idx) in covered:
                continue
            cell = ws.cell(row_idx, col_idx)
            merge = merged_lookup.get((row_idx, col_idx))
            row_cells.append({
                "col": col_idx,
                "value": normalize_cell(cell.value),
                "rowspan": merge["rowspan"] if merge else 1,
                "colspan": merge["colspan"] if merge else 1,
            })
        rows.append({"row": row_idx, "cells": row_cells})

    column_widths = []
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        width = ws.column_dimensions[letter].width
        column_widths.append(width if width else 13)

    return {
        "source_file": TEMPLATE_2024.name,
        "sheet_name": ws.title,
        "max_row": ws.max_row,
        "max_col": max_col,
        "column_widths": column_widths,
        "rows": rows,
    }


# ────────────────────────────────────────────────
# Docling API による PDF テキスト抽出
# ────────────────────────────────────────────────

BACKEND_PRIORITY = ["dlparse_v2", "dlparse_v4"]
MIN_ENTRIES_THRESHOLD = 1  # 1件以上取れたら成功とする（複数バックエンド試行で Docling を疲弊させない）


def _call_docling(pdf_path: Path, backend: str, timeout: int) -> str:
    """指定バックエンドで Docling API を呼び出し Markdown を返す。"""
    with pdf_path.open("rb") as f:
        resp = requests.post(
            DOCLING_URL,
            files={"files": (pdf_path.name, f, "application/pdf")},
            data={"options": f'{{"to_formats":["md"],"pdf_backend":"{backend}","do_ocr":false}}'},
            timeout=timeout,
        )
    resp.raise_for_status()
    result = resp.json()
    docs = result.get("document", result.get("documents", []))
    if isinstance(docs, dict):
        return docs.get("md_content", "")
    if isinstance(docs, list) and docs:
        return docs[0].get("md_content", "")
    return ""


DOCLING_CONTAINER = "clawstack-unified-docling-1"
DOCLING_RESTART_WAIT = 25  # コンテナ起動完了を待つ秒数


def _restart_docling() -> None:
    """Docling コンテナを再起動して詰まったジョブをクリアする。"""
    try:
        subprocess.run(
            ["docker", "restart", DOCLING_CONTAINER],
            check=True, capture_output=True, timeout=30,
        )
        print(f"    Docling restarted, waiting {DOCLING_RESTART_WAIT}s ...", flush=True)
        time.sleep(DOCLING_RESTART_WAIT)
    except Exception as exc:
        print(f"    WARNING: could not restart Docling: {exc}", flush=True)
        time.sleep(30)  # フォールバック: 待つだけ


DOCLING_CLIENT_TIMEOUT = 320  # max_sync_wait=300 + バッファ


def extract_pdf_via_docling(pdf_path: Path, timeout: int = DOCLING_CLIENT_TIMEOUT) -> str:
    """複数バックエンドを試し、エントリ数が最大のものを返す。

    504 / ReadTimeout は Docling がタイムアウトしたシグナル。
    この場合は即座に諦め、Docling を再起動して次の PDF に進む。
    成功したが entries 数が少ない場合のみ次のバックエンドを試す。
    """
    best_md = ""
    best_count = 0

    for backend in BACKEND_PRIORITY:
        try:
            print(f"    backend={backend} ...", flush=True)
            md = _call_docling(pdf_path, backend, timeout)
            count = len(parse_docling_entries(md))
            print(f"    → {count} entries", flush=True)
            if count > best_count:
                best_count = count
                best_md = md
            if count >= MIN_ENTRIES_THRESHOLD:
                break  # 十分な件数が取れたら終了
        except requests.HTTPError as exc:
            status = exc.response.status_code if exc.response is not None else 0
            print(f"    backend={backend} HTTP {status}: {exc}", flush=True)
            if status == 504:
                print("    504 Timeout: restarting Docling to clear stuck job ...", flush=True)
                _restart_docling()
                break  # 504 の場合は他のバックエンドも試さない
            continue
        except requests.exceptions.Timeout as exc:
            # クライアント側タイムアウト: Docling がまだ処理中 → 再起動
            print(f"    backend={backend} Client Timeout: restarting Docling ...", flush=True)
            _restart_docling()
            break
        except Exception as exc:
            print(f"    backend={backend} ERROR: {exc}", flush=True)
            continue

    return best_md


# ────────────────────────────────────────────────
# Markdown テーブル → entries 変換
# ────────────────────────────────────────────────

def parse_md_table_row(line: str) -> list[str]:
    """Markdown テーブル行をセルリストに変換する。"""
    cells = [c.strip() for c in line.split("|")]
    # 前後の空要素を除去
    if cells and cells[0] == "":
        cells = cells[1:]
    if cells and cells[-1] == "":
        cells = cells[:-1]
    return cells


def is_target_like(text: str) -> bool:
    return bool(TARGET_RE.search(text)) if text else False


def is_actual_like(text: str) -> bool:
    return bool(ACTUAL_RE.match(text)) if text else False


def looks_like_process(text: str) -> bool:
    return any(kw in text for kw in PROCESS_KEYWORDS)


def parse_docling_entries(md_content: str) -> list[dict]:
    """Docling の Markdown テーブルを entries リストに変換する。

    テーブル形式:
      5列: | 区分 | プロセス | 指標 | 目標 | 実績 |  (2024年8月など)
      4列: | プロセス | 指標 | 目標 | 実績 |      (2025年10月・11月など)
    """
    entries: list[dict] = []
    current_process = ""
    in_table = False
    header_skipped = False
    table_ncols = 0  # ヘッダ行で確定した列数

    for line in md_content.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            if in_table:
                in_table = False
                header_skipped = False
                table_ncols = 0
            continue

        in_table = True
        cells = parse_md_table_row(line)

        # セパレータ行でヘッダ列数を確定
        if all(re.fullmatch(r"[-: ]+", c) for c in cells if c):
            header_skipped = True
            continue
        if not header_skipped:
            # ヘッダ行: 列数を記録
            if len(cells) >= 4:
                table_ncols = len(cells)
            continue
        if len(cells) < 4:
            continue

        ncols = table_ncols if table_ncols else len(cells)

        if ncols >= 5:
            # ─── 5列形式 ───────────────────────────────────────
            # | 区分 | プロセス | 指標 | 目標 | 実績 |
            c0 = cells[0] if len(cells) > 0 else ""
            c1 = cells[1] if len(cells) > 1 else ""
            c2 = cells[2] if len(cells) > 2 else ""
            c3 = cells[3] if len(cells) > 3 else ""
            c4 = cells[4] if len(cells) > 4 else ""

            has_target = is_target_like(c3) or is_target_like(c2)
            has_actual = is_actual_like(c4) or is_actual_like(c3)
            if not (has_target or has_actual):
                continue

            if c2:
                metric = c2
                target = c3
                actual = c4
                if c1 and looks_like_process(c1):
                    current_process = c1
            else:
                metric = c1
                target = c3
                actual = c4
        else:
            # ─── 4列形式 ───────────────────────────────────────
            # | プロセス | 指標 | 目標 | 実績 |
            c0 = cells[0] if len(cells) > 0 else ""
            c1 = cells[1] if len(cells) > 1 else ""
            c2 = cells[2] if len(cells) > 2 else ""
            c3 = cells[3] if len(cells) > 3 else ""

            has_target = is_target_like(c2)
            has_actual = is_actual_like(c3) or is_actual_like(c2)
            if not (has_target or has_actual):
                continue

            # c0 がプロセス名の場合は更新（セクション記号・日付注記は除外）
            if c0 and looks_like_process(c0):
                current_process = c0
            elif c0 and not re.match(r"^[ⅠⅡⅢⅣⅤⅵⅶ・]", c0) and c0 not in {"", "対", "応 方", "見", "等", "管 理"}:
                current_process = c0

            metric = c1
            target = c2
            actual = c3

        if not metric or not (target or actual):
            continue

        entries.append({
            "process": current_process,
            "metric": metric,
            "target": target,
            "actual": actual,
        })

    return entries


# ────────────────────────────────────────────────
# PDF 1件を処理
# ────────────────────────────────────────────────

def extract_year_pdf(pdf_path: Path, default_year: int) -> dict | None:
    print(f"  Docling: {pdf_path.name} ...", flush=True)
    try:
        md_content = extract_pdf_via_docling(pdf_path)
    except Exception as exc:
        print(f"    ERROR: {exc}", file=sys.stderr, flush=True)
        return None

    month_match = MONTH_PATTERN.search(pdf_path.name) or MONTH_PATTERN.search(md_content)
    created_match = DATE_PATTERN.search(md_content)

    entries = parse_docling_entries(md_content)
    print(f"    → final: {len(entries)} entries", flush=True)

    return {
        "source_file": pdf_path.name,
        "year": int(month_match.group(1)) if month_match else default_year,
        "month": int(month_match.group(2)) if month_match else None,
        "created_date": created_match.group(1) if created_match else "",
        "entries": entries,
        "raw_preview": md_content[:2000],
    }


def extract_year(year: int) -> list[dict]:
    year_dir = PDF_YEAR_DIRS.get(year)
    if not year_dir:
        return []

    files = sorted(
        year_dir.glob("*.pdf"),
        key=lambda p: int(MONTH_PATTERN.search(p.name).group(2)) if MONTH_PATTERN.search(p.name) else 99,
    )
    results = []
    for i, pdf in enumerate(files):
        if i > 0:
            time.sleep(10)  # Docling が前リクエストから回復する時間を与える
        item = extract_year_pdf(pdf, year)
        if item is not None:
            results.append(item)
    return results


# ────────────────────────────────────────────────
# メイン
# ────────────────────────────────────────────────

def main():
    import datetime
    print("=== extract_process_monitoring_measurement ===")
    print(f"SOURCE_DIR: {SOURCE_DIR}")

    payload = {
        "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "source_dir": str(SOURCE_DIR),
        "year_2024": extract_2024(),
    }
    print("2024: Excel OK")

    for year in sorted(PDF_YEAR_DIRS):
        print(f"{year}: processing PDFs...")
        payload[f"year_{year}"] = extract_year(year)

    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Written: {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
