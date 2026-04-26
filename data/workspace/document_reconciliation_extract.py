from __future__ import annotations

import hashlib
import json
import re
import shutil
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

import openpyxl
import xlrd


ROOT = Path(r"D:\Clawdbot_Docker_20260125")
SOURCE_ZIP = ROOT / "clawstack_v2" / "data" / "paperless" / "consume" / "文書照合_20250329.zip"
EXTRACT_DIR = ROOT / "data" / "workspace" / "_tmp_doc_reconcile"
REPORT_PATH = ROOT / "iatf_system" / "db" / "document_reconciliation_report.json"


QM_FILE_NAME = "QM001付表4_文書（記録を含む）体系表_第5版_20250110.xlsx"
RECORD_RULE_FILE_NAME = "記録管理規定 表1　25.1.10改.xlsx"
NUMBERING_DOC_FILE_NAME = "記録の採番一覧.doc"

DEPARTMENT_FILES = {
    "設計部門": ["設計部門_管理文書一覧.xls"],
    "生産技術": ["生産技術管理文書一覧.xls", "第三階層文書・文書見直し記録(KGO007)_2025_1_20.xls"],
    "プレス_表面処理": ["プレス_処理部門_ISO管理文書リスト（番号・改訂日・レビュー）.xls"],
    "成形": ["第三階層文書・文書見直し記録(KSOa903) 2021.11.xls", "第二階層文書管理台帳（KSOa906)（文書見直し記録）2022.11.xls"],
    "営業部": ["営業部管理文書台帳EG-008-9　(最新）.xls"],
    "管理部": ["管理部　文書管理台帳.xls"],
}

DEPARTMENT_PREFIXES = {
    "設計部門": ("KGS", "GTS"),
    "生産技術": ("KGI", "KGO", "GTI", "GTO", "GTP"),
    "プレス_表面処理": ("SZ", "KSS", "HH"),
    "成形": ("AK", "KSO", "SZR"),
    "営業部": ("EG", "KE"),
    "管理部": ("GM", "KGM", "GMA"),
}

DOC_NUMBER_RE = re.compile(r"[A-Z]{1,5}[0-9]{2,4}(?:[-‐－ー][A-Z0-9]{1,8})*")


@dataclass
class Entry:
    source: str
    source_file: str
    sheet: str
    entry_type: str
    name: str
    number: str
    retention: str
    department: str

    @property
    def normalized_name(self) -> str:
        return normalize_text(self.name)

    @property
    def family_name(self) -> str:
        return family_name(self.name)

    @property
    def normalized_number(self) -> str:
        return normalize_number(self.number)


def normalize_text(value: object) -> str:
    text = str(value or "").replace("\u3000", " ").replace("\n", " ").replace("\r", " ")
    text = text.replace("（", "(").replace("）", ")")
    text = text.replace("・", "").replace("･", "")
    return re.sub(r"\s+", "", text).strip().upper()


def family_name(value: object) -> str:
    text = str(value or "").strip()
    text = text.replace("（", "(").replace("）", ")")
    text = text.replace("Ｄ", "D").replace("Ｒ", "R").replace("・", "・")
    rules = [
        ("設計計画書", "設計計画書"),
        ("D・R会議議事録", "DR会議議事録"),
        ("DR会議議事録", "DR会議議事録"),
        ("D.R会議議事録", "DR会議議事録"),
        ("設計検証チェックリスト", "設計検証チェックリスト"),
        ("設計変更会議議事録", "設計変更会議議事録"),
        ("DR構想検討会議議事録", "DR構想検討会議議事録"),
    ]
    for pattern, canonical in rules:
        if pattern in text:
            return canonical

    text = re.sub(r"[\(（].*?[\)）]", "", text)
    text = re.sub(r"\s+", "", text)
    return text


def normalize_number(value: object) -> str:
    text = str(value or "").strip()
    text = text.replace("\n", "").replace("\r", "").replace("\u3000", "")
    text = text.replace("‐", "-").replace("－", "-").replace("ー", "-")
    text = text.replace("―", "-")
    return text.upper()


def retention_year_token(value: str) -> str:
    match = re.search(r"(\d+)", str(value or ""))
    return match.group(1) if match else ""


def extract_doc_number(value: object) -> str:
    text = str(value or "")
    match = DOC_NUMBER_RE.search(text.upper().replace("（", "(").replace("）", ")"))
    return normalize_number(match.group(0)) if match else ""


def classify_department(number: str) -> str:
    normalized = normalize_number(number)
    if not normalized:
        return ""
    for department, prefixes in DEPARTMENT_PREFIXES.items():
        if normalized.startswith(prefixes):
            return department
    return ""


def is_blank_row(values: Iterable[object]) -> bool:
    return not any(str(v or "").strip() for v in values)


def workbook_rows(path: Path) -> list[tuple[str, list[list[object]]]]:
    if path.suffix.lower() == ".xlsx":
        wb = openpyxl.load_workbook(path, data_only=True)
        return [
            (ws.title, [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(1, ws.max_row + 1)])
            for ws in wb.worksheets
        ]

    wb = xlrd.open_workbook(path)
    rows = []
    for sheet_name in wb.sheet_names():
        sheet = wb.sheet_by_name(sheet_name)
        rows.append((sheet_name, [sheet.row_values(r, 0, sheet.ncols) for r in range(sheet.nrows)]))
    return rows


def prepare_extract_dir() -> Path:
    if EXTRACT_DIR.exists():
        shutil.rmtree(EXTRACT_DIR)
    EXTRACT_DIR.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(SOURCE_ZIP) as zf:
        zf.extractall(EXTRACT_DIR)
    matches = list(EXTRACT_DIR.glob("文書照合_*"))
    if not matches:
        raise FileNotFoundError("文書照合 ZIP の展開先が見つかりません。")
    return matches[0]


def parse_qm_entries(base_dir: Path) -> list[Entry]:
    path = base_dir / QM_FILE_NAME
    rows = workbook_rows(path)[0][1]
    entries: list[Entry] = []
    for row in rows:
        row = list(row) + [None] * max(0, 15 - len(row))
        record_name = str(row[7] or "").strip()
        if record_name and record_name not in {"記録", "品質記録"}:
            number = normalize_number(row[8])
            entries.append(
                Entry(
                    source="qm001_appendix4",
                    source_file=path.name,
                    sheet="QM001付表4",
                    entry_type="record",
                    name=record_name,
                    number=number,
                    retention=str(row[13] or "").strip(),
                    department=classify_department(number),
                )
            )

        for col in (3, 4, 5):
            doc_name = str(row[col] or "").strip()
            if not doc_name or "文書" == doc_name or "規定" == doc_name:
                continue
            number = extract_doc_number(doc_name)
            entries.append(
                Entry(
                    source="qm001_appendix4",
                    source_file=path.name,
                    sheet="QM001付表4",
                    entry_type="document",
                    name=doc_name,
                    number=number,
                    retention="",
                    department=classify_department(number),
                )
            )
    return dedupe_entries(entries)


def parse_record_rule_entries(base_dir: Path) -> list[Entry]:
    path = base_dir / RECORD_RULE_FILE_NAME
    all_rows = {sheet: rows for sheet, rows in workbook_rows(path)}
    rows = all_rows.get("Sheet2") or next(iter(all_rows.values()))
    entries: list[Entry] = []
    for row in rows:
        row = list(row) + [None] * max(0, 18 - len(row))
        name = str(row[3] or "").strip()
        if not name or name in {"品質記録", "文書（記録を含む）体系表"}:
            continue
        entries.append(
            Entry(
                source="record_rule_table",
                source_file=path.name,
                sheet="Sheet2",
                entry_type="record_rule",
                name=name,
                number=normalize_number(row[8]),
                retention=str(row[17] or "").strip(),
                department=classify_department(row[8]),
            )
        )
    return dedupe_entries(entries)


def find_header_indices(header: list[object]) -> tuple[int | None, int | None, int | None]:
    name_idx = None
    number_idx = None
    retention_idx = None
    for idx, value in enumerate(header):
        text = str(value or "").strip()
        if not text:
            continue
        if name_idx is None and any(token in text for token in ["文書名", "記録書名", "品質記録", "記録表", "記録名"]):
            name_idx = idx
        if number_idx is None and any(token in text for token in ["文書番号", "管理番号", "記録様式番号", "新文書番号"]):
            number_idx = idx
        if retention_idx is None and "保管期間" in text:
            retention_idx = idx
    if number_idx is not None and name_idx is None:
        name_idx = 0
    return name_idx, number_idx, retention_idx


def parse_department_entries(base_dir: Path) -> dict[str, list[Entry]]:
    results: dict[str, dict[str, object]] = {}
    for department, file_names in DEPARTMENT_FILES.items():
        entries: list[Entry] = []
        adopted_sheets: list[dict[str, object]] = []
        inspected_files: list[dict[str, object]] = []
        for file_name in file_names:
            path = next(base_dir.rglob(file_name))
            workbook = workbook_rows(path)
            inspected_files.append(
                {
                    "source_file": path.name,
                    "sheet_names": [sheet_name for sheet_name, _ in workbook],
                }
            )
            for sheet_name, rows in workbook:
                header_index = None
                header_info = (None, None, None)
                for idx, row in enumerate(rows[:30]):
                    name_idx, number_idx, retention_idx = find_header_indices(row)
                    if name_idx is not None and number_idx is not None:
                        header_index = idx
                        header_info = (name_idx, number_idx, retention_idx)
                        break
                if header_index is None:
                    continue

                name_idx, number_idx, retention_idx = header_info
                before_count = len(entries)
                blank_streak = 0
                for row in rows[header_index + 1 :]:
                    row = list(row)
                    if is_blank_row(row):
                        blank_streak += 1
                        if blank_streak >= 6:
                            break
                        continue
                    blank_streak = 0

                    name = str(row[name_idx] if name_idx < len(row) else "").strip()
                    number = normalize_number(row[number_idx] if number_idx < len(row) else "")
                    retention = str(row[retention_idx] if retention_idx is not None and retention_idx < len(row) else "").strip()

                    if not name and not number:
                        continue
                    if name in {"文書名", "記録書名", "品質記録"} or number in {"管理番号", "文書番号", "新文書番号"}:
                        continue
                    if not name or len(normalize_text(name)) <= 1:
                        continue

                    entries.append(
                        Entry(
                            source="department_ledger",
                            source_file=path.name,
                            sheet=sheet_name,
                            entry_type="ledger",
                            name=name,
                            number=number,
                            retention=retention,
                            department=department,
                        )
                    )
                sheet_entry_count = len(entries) - before_count
                if sheet_entry_count > 0:
                    adopted_sheets.append(
                        {
                            "source_file": path.name,
                            "sheet": sheet_name,
                            "entry_count": sheet_entry_count,
                            "header_row": header_index + 1,
                            "name_column_index": name_idx + 1 if name_idx is not None else None,
                            "number_column_index": number_idx + 1 if number_idx is not None else None,
                            "retention_column_index": retention_idx + 1 if retention_idx is not None else None,
                        }
                    )
        results[department] = {
            "entries": dedupe_entries(entries),
            "adopted_sheets": adopted_sheets,
            "inspected_files": inspected_files,
        }
    return results


def dedupe_entries(entries: list[Entry]) -> list[Entry]:
    seen = set()
    unique: list[Entry] = []
    for entry in entries:
        key = (entry.source, entry.entry_type, entry.department, entry.family_name, entry.normalized_name, entry.normalized_number)
        if key in seen:
            continue
        seen.add(key)
        unique.append(entry)
    return unique


def numbers_compatible(left: str, right: str) -> bool:
    lnorm = normalize_number(left)
    rnorm = normalize_number(right)
    if not lnorm or not rnorm:
        return True
    if lnorm == rnorm:
        return True
    shorter, longer = sorted([lnorm, rnorm], key=len)
    return len(shorter) <= 4 and longer.startswith(shorter)


def find_best_match(target: Entry, candidates: list[Entry]) -> Entry | None:
    exact = [candidate for candidate in candidates if candidate.normalized_name == target.normalized_name]
    if exact:
        return exact[0]
    family = [candidate for candidate in candidates if candidate.family_name == target.family_name]
    if family:
        for candidate in family:
            if numbers_compatible(target.number, candidate.number):
                return candidate
        return family[0]
    return None


def build_core_mismatches(qm_entries: list[Entry], rule_entries: list[Entry]) -> list[dict]:
    qm_records = [entry for entry in qm_entries if entry.entry_type == "record"]
    rule_by_family = defaultdict(list)
    for entry in rule_entries:
        rule_by_family[entry.family_name].append(entry)

    mismatches = []
    matched_families = set()
    for qm_entry in qm_records:
        candidates = rule_by_family.get(qm_entry.family_name, [])
        if not candidates:
            mismatches.append(
                {
                    "type": "missing_in_record_rule",
                    "family_name": qm_entry.family_name,
                    "qm_name": qm_entry.name,
                    "qm_number": qm_entry.number,
                    "qm_retention": qm_entry.retention,
                    "rule_name": "",
                    "rule_number": "",
                    "rule_retention": "",
                }
            )
            continue

        matched_families.add(qm_entry.family_name)
        rule_entry = candidates[0]
        if not numbers_compatible(qm_entry.number, rule_entry.number):
            mismatches.append(
                {
                    "type": "number_mismatch",
                    "family_name": qm_entry.family_name,
                    "qm_name": qm_entry.name,
                    "qm_number": qm_entry.number,
                    "qm_retention": qm_entry.retention,
                    "rule_name": rule_entry.name,
                    "rule_number": rule_entry.number,
                    "rule_retention": rule_entry.retention,
                }
            )
        if retention_year_token(qm_entry.retention) != retention_year_token(rule_entry.retention):
            mismatches.append(
                {
                    "type": "retention_mismatch",
                    "family_name": qm_entry.family_name,
                    "qm_name": qm_entry.name,
                    "qm_number": qm_entry.number,
                    "qm_retention": qm_entry.retention,
                    "rule_name": rule_entry.name,
                    "rule_number": rule_entry.number,
                    "rule_retention": rule_entry.retention,
                }
            )

    for rule_entry in rule_entries:
        if rule_entry.family_name not in matched_families:
            mismatches.append(
                {
                    "type": "missing_in_qm001",
                    "family_name": rule_entry.family_name,
                    "qm_name": "",
                    "qm_number": "",
                    "qm_retention": "",
                    "rule_name": rule_entry.name,
                    "rule_number": rule_entry.number,
                    "rule_retention": rule_entry.retention,
                }
            )

    return mismatches


def build_department_mismatches(qm_entries: list[Entry], department_entries: dict[str, list[Entry]]) -> list[dict]:
    results = []
    for department, payload in department_entries.items():
        ledger_entries = payload["entries"]
        applicable_qm = [entry for entry in qm_entries if entry.department == department]
        matched_qm_keys = set()
        matched_ledger_keys = set()
        number_mismatches = []
        missing_in_baseline = []

        for ledger_entry in ledger_entries:
            matched = find_best_match(ledger_entry, applicable_qm)
            if matched is None:
                missing_in_baseline.append(
                    {
                        "name": ledger_entry.name,
                        "number": ledger_entry.number,
                        "sheet": ledger_entry.sheet,
                        "source_file": ledger_entry.source_file,
                    }
                )
                continue

            matched_qm_keys.add((matched.family_name, matched.normalized_name, matched.normalized_number))
            matched_ledger_keys.add((ledger_entry.family_name, ledger_entry.normalized_name, ledger_entry.normalized_number))
            if not numbers_compatible(ledger_entry.number, matched.number):
                number_mismatches.append(
                    {
                        "baseline_name": matched.name,
                        "baseline_number": matched.number,
                        "ledger_name": ledger_entry.name,
                        "ledger_number": ledger_entry.number,
                        "sheet": ledger_entry.sheet,
                        "source_file": ledger_entry.source_file,
                    }
                )

        missing_in_department = []
        for qm_entry in applicable_qm:
            key = (qm_entry.family_name, qm_entry.normalized_name, qm_entry.normalized_number)
            if key not in matched_qm_keys:
                missing_in_department.append(
                    {
                        "name": qm_entry.name,
                        "number": qm_entry.number,
                        "retention": qm_entry.retention,
                        "source_file": qm_entry.source_file,
                    }
                )

        results.append(
            {
                "department": department,
                "baseline_count": len(applicable_qm),
                "ledger_count": len(ledger_entries),
                "adopted_sheets": payload["adopted_sheets"],
                "inspected_files": payload["inspected_files"],
                "missing_in_department": sorted(missing_in_department, key=lambda item: (item["name"], item["number"])),
                "missing_in_baseline": sorted(missing_in_baseline, key=lambda item: (item["name"], item["number"])),
                "number_mismatches": sorted(number_mismatches, key=lambda item: (item["ledger_name"], item["ledger_number"])),
            }
        )

    return results


def file_meta(path: Path) -> dict:
    stat = path.stat()
    return {
        "name": path.name,
        "relative_path": str(path.relative_to(ROOT)).replace("\\", "/"),
        "size_bytes": stat.st_size,
        "sha1": hashlib.sha1(path.read_bytes()).hexdigest(),
        "updated_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
    }


def numbering_doc_status(base_dir: Path) -> dict:
    path = base_dir / NUMBERING_DOC_FILE_NAME
    return {
        "file": file_meta(path),
        "status": "unparsed",
        "reason": "この環境では Word COM / antiword / LibreOffice が利用できず、.doc 本文を自動抽出できませんでした。存在確認とハッシュ取得のみ実施しています。",
    }


def main() -> None:
    base_dir = prepare_extract_dir()
    qm_entries = parse_qm_entries(base_dir)
    rule_entries = parse_record_rule_entries(base_dir)
    department_entries = parse_department_entries(base_dir)
    core_mismatches = build_core_mismatches(qm_entries, rule_entries)
    department_mismatches = build_department_mismatches(qm_entries, department_entries)

    report = {
        "generated_at": datetime.now().isoformat(),
        "source_zip": file_meta(SOURCE_ZIP),
        "numbering_reference": numbering_doc_status(base_dir),
        "source_files": {
            "qm001_appendix4": file_meta(base_dir / QM_FILE_NAME),
            "record_rule_table": file_meta(base_dir / RECORD_RULE_FILE_NAME),
            "department_ledgers": {
                department: [file_meta(next(base_dir.rglob(file_name))) for file_name in file_names]
                for department, file_names in DEPARTMENT_FILES.items()
            },
        },
        "summary": {
            "qm_entry_count": len(qm_entries),
            "record_rule_count": len(rule_entries),
            "department_entry_count": sum(len(payload["entries"]) for payload in department_entries.values()),
            "core_mismatch_count": len(core_mismatches),
            "department_mismatch_count": sum(
                len(item["missing_in_department"]) + len(item["missing_in_baseline"]) + len(item["number_mismatches"])
                for item in department_mismatches
            ),
        },
        "core_mismatches": core_mismatches,
        "department_mismatches": department_mismatches,
    }

    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(str(REPORT_PATH))


if __name__ == "__main__":
    main()
