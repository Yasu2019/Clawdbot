# -*- coding: utf-8 -*-
"""
excel_to_txt_for_paperless.py

Converts Excel files in Paperless consume directories to .txt files
so Paperless can ingest their text content.
Supports .xlsx, .xls, .xlsm

Usage:
  python scripts/excel_to_txt_for_paperless.py
"""
import io
import sys
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

CONSUME_ROOT = Path(r"D:\Clawdbot_Docker_20260125\clawstack_v2\data\paperless\consume")

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    print("[WARN] openpyxl not installed — install with: pip install openpyxl")

try:
    import xlrd
    XLRD_OK = True
except ImportError:
    XLRD_OK = False


def excel_to_text_openpyxl(path: Path) -> str:
    """Extract text from .xlsx/.xlsm using openpyxl."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    lines = [f"[FILE] {path.name}"]
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        lines.append(f"\n[SHEET] {sheet}")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                lines.append("\t".join(cells))
    wb.close()
    return "\n".join(lines)


def excel_to_text_xlrd(path: Path) -> str:
    """Extract text from .xls using xlrd."""
    wb = xlrd.open_workbook(str(path))
    lines = [f"[FILE] {path.name}"]
    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        lines.append(f"\n[SHEET] {sheet_name}")
        for r in range(ws.nrows):
            cells = [str(ws.cell_value(r, c)).strip()
                     for c in range(ws.ncols)
                     if str(ws.cell_value(r, c)).strip()]
            if cells:
                lines.append("\t".join(cells))
    return "\n".join(lines)


def convert_excel(path: Path) -> bool:
    ext = path.suffix.lower()
    out = path.with_suffix(".txt")
    if out.exists():
        return False  # already done

    if path.stat().st_size == 0:
        return False  # empty stub file

    try:
        if ext in (".xlsx", ".xlsm") and OPENPYXL_OK:
            text = excel_to_text_openpyxl(path)
        elif ext == ".xls" and XLRD_OK:
            text = excel_to_text_xlrd(path)
        elif ext in (".xlsx", ".xlsm") and not OPENPYXL_OK:
            print(f"  [SKIP] openpyxl missing: {path.name}")
            return False
        elif ext == ".xls" and not XLRD_OK:
            print(f"  [SKIP] xlrd missing: {path.name}")
            return False
        else:
            return False

        out.write_text(text, encoding="utf-8")
        print(f"  [OK] {path.name} → {out.name} ({len(text):,} chars)")
        return True
    except Exception as e:
        print(f"  [ERR] {path.name}: {e}")
        return False


def main():
    print(f"[START] Excel → txt converter")
    print(f"[INFO]  Scanning: {CONSUME_ROOT}")
    print()

    total = 0
    converted = 0

    for ext in ("*.xlsx", "*.xls", "*.xlsm"):
        for path in CONSUME_ROOT.rglob(ext):
            total += 1
            if convert_excel(path):
                converted += 1

    print()
    print(f"=== 完了 ===")
    print(f"Excel ファイル検出: {total}")
    print(f"テキスト変換成功: {converted}")
    print(f"スキップ(変換済み): {total - converted}")
    print(f"変換後 .txt は Paperless が自動取り込みします")


if __name__ == "__main__":
    main()
