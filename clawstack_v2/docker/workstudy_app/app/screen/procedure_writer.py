"""
ProcedureWriter — Generates two output files from screen recording analysis:

  1. procedure_text.txt
       Plain-text log of every detected operation with OCR-extracted content.
       Phase 1 — "テキスト文書の抽出"

  2. procedure.xlsx
       Step-by-step Excel work instruction (手順書) with:
         • Embedded screenshot thumbnail per click event
         • Red border (赤枠) around the exact click area in each thumbnail
         • OCR text extracted near each click
         • Formatted table: Step / 時刻 / スクリーンショット / 操作内容 / 座標 / 備考
       Phase 2 — "画像付き、赤枠付き手順書"
"""

import io
import cv2
import numpy as np
from pathlib import Path
from datetime import datetime

from PIL import Image, ImageDraw, ImageFont

import openpyxl
from openpyxl.styles import Font as XFont, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from screen.ocr_extractor import OCRExtractor


# ── Excel styling constants ───────────────────────────────────────────────────
_FILL_TITLE  = PatternFill("solid", fgColor="1E3A5F")   # dark navy
_FILL_HEAD   = PatternFill("solid", fgColor="2563EB")   # Clawstack blue
_FILL_EVEN   = PatternFill("solid", fgColor="EBF5FF")   # light blue stripe
_FILL_ODD    = PatternFill("solid", fgColor="FFFFFF")
_FILL_WARN   = PatternFill("solid", fgColor="FFF3CD")   # amber for dblclick

_SIDE        = Side(style="thin", color="BBBBBB")
_BORDER      = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)

# Thumbnail dimensions written into the Excel cells (in pixels)
_THUMB_W    = 320
_THUMB_H    = 200
# Red-border box margin around click point (pixels in *original* resolution)
_BOX_MARGIN = 70
# Excel row height for a thumbnail row (points; 1 px ≈ 0.75 pt)
_ROW_HEIGHT = _THUMB_H * 0.75 + 8


class ProcedureWriter:
    """Generates text log + Excel procedure from click events and a screen recording."""

    # ── Public entry points ───────────────────────────────────────────────────

    def write_text_log(
        self,
        events: list[dict],
        video_path: str,
        project_dir: Path,
        title: str = "PC操作 作業手順書",
        progress_cb=None,
    ) -> Path:
        """
        Phase 1: Generate plain-text procedure log with OCR content.

        Returns:
            Path to procedure_text.txt
        """
        txt_path = project_dir / "procedure_text.txt"
        ocr = OCRExtractor()
        ocr_available = OCRExtractor.is_available()

        cap = cv2.VideoCapture(video_path)
        fps = cap.get(cv2.CAP_PROP_FPS) or 30.0

        lines = [
            f"{'=' * 60}",
            f"  {title}",
            f"  生成日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"  録画ファイル: {Path(video_path).name}",
            f"  OCR: {'有効 (Tesseract)' if ocr_available else '無効 (Tesseract未インストール)'}",
            f"{'=' * 60}",
            "",
        ]

        total = max(len(events), 1)
        for step_no, ev in enumerate(events, start=1):
            # Real-time progress
            if progress_cb:
                progress_cb(
                    step_no / total,
                    f"OCRテキスト抽出中... Step {step_no}/{total}",
                )

            ts_sec = ev["time_sec"]
            mm, ss = divmod(int(ts_sec), 60)
            kind   = "ダブルクリック" if ev.get("type") == "dblclick" else "クリック"

            lines.append(f"【Step {step_no:03d}】 {mm:02d}:{ss:02d}  {kind}  座標({ev['x']}, {ev['y']})")

            # OCR text near click
            if ocr_available:
                cap.set(cv2.CAP_PROP_POS_FRAMES, ev["frame"])
                ret, frame = cap.read()
                if ret:
                    # Wider extraction (400px) for full context
                    text = ocr.extract_near_click(frame, ev["x"], ev["y"], radius=400)
                    if text:
                        lines.append("  ▼ 画面テキスト:")
                        for tline in text.split("\n"):
                            tline = tline.strip()
                            if tline:
                                lines.append(f"    {tline}")
                    else:
                        lines.append("  ▼ 画面テキスト: (検出なし)")

            lines.append("")

        cap.release()

        with open(txt_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

        return txt_path

    def write_excel(
        self,
        events: list[dict],
        video_path: str,
        project_dir: Path,
        title: str = "PC操作 作業手順書",
        progress_cb=None,
    ) -> Path:
        """
        Phase 2: Generate Excel work instruction with screenshots + 赤枠 + OCR text.

        Returns:
            Path to procedure.xlsx
        """
        xlsx_path = project_dir / "procedure.xlsx"
        ocr = OCRExtractor()
        ocr_available = OCRExtractor.is_available()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "手順書"

        # ── Sheet setup ──────────────────────────────────────────────────────
        self._write_title_row(ws, title)
        self._write_col_headers(ws)

        # ── Column widths ────────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 8    # Step
        ws.column_dimensions["B"].width = 10   # 時刻
        ws.column_dimensions["C"].width = 46   # Screenshot (≈320px / 7px per char)
        ws.column_dimensions["D"].width = 45   # 操作内容 (OCR)
        ws.column_dimensions["E"].width = 16   # 座標
        ws.column_dimensions["F"].width = 22   # 備考

        # ── Data rows ────────────────────────────────────────────────────────
        cap   = cv2.VideoCapture(video_path)
        fps   = cap.get(cv2.CAP_PROP_FPS) or 30.0
        total = max(len(events), 1)

        data_row = 3
        for step_no, ev in enumerate(events, start=1):
            # Real-time progress
            if progress_cb:
                progress_cb(
                    step_no / total,
                    f"手順書生成中... Step {step_no}/{total}  スクリーンショット取得 + 赤枠描画",
                )

            # Seek to click frame
            cap.set(cv2.CAP_PROP_POS_FRAMES, ev["frame"])
            ret, frame = cap.read()
            if not ret:
                continue

            # OCR near click
            ocr_text = ""
            if ocr_available:
                ocr_text = ocr.extract_near_click(frame, ev["x"], ev["y"], radius=220)

            # Annotated thumbnail (PNG bytes)
            thumb_png = self._make_thumbnail(frame, ev["x"], ev["y"])

            # Write the row
            self._write_step_row(
                ws, step_no, ev, fps, ocr_text, thumb_png, data_row
            )
            data_row += 1

        cap.release()

        # ── Click log sheet ──────────────────────────────────────────────────
        ws_log = wb.create_sheet("クリックログ")
        self._write_log_sheet(ws_log, events, fps)

        # ── OCR summary sheet (if OCR available) ─────────────────────────────
        # (already captured in 手順書; skip separate sheet to keep workbook lean)

        wb.save(str(xlsx_path))
        return xlsx_path

    # ── Excel writing helpers ─────────────────────────────────────────────────

    def _write_title_row(self, ws, title: str):
        ws.merge_cells("A1:F1")
        cell = ws["A1"]
        cell.value     = title
        cell.font      = XFont(name="メイリオ", size=16, bold=True, color="FFFFFF")
        cell.fill      = _FILL_TITLE
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 34

    def _write_col_headers(self, ws):
        headers = ["Step", "時刻", "スクリーンショット（赤枠=クリック箇所）",
                   "操作内容 / 画面テキスト (OCR)", "座標 (X, Y)", "備考"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font      = XFont(name="メイリオ", size=10, bold=True, color="FFFFFF")
            cell.fill      = _FILL_HEAD
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border    = _BORDER
        ws.row_dimensions[2].height = 24

    def _write_step_row(self, ws, step_no, ev, fps, ocr_text, thumb_png, row):
        ts_sec    = ev["time_sec"]
        mm, ss    = divmod(int(ts_sec), 60)
        ts_str    = f"{mm:02d}:{ss:02d}"
        kind      = "ダブルクリック" if ev.get("type") == "dblclick" else "クリック"
        coord_str = f"({ev['x']}, {ev['y']})"

        # Description: kind + OCR text
        desc = f"[{kind}]\n{ocr_text}" if ocr_text else f"[{kind}]"

        fill = _FILL_WARN if ev.get("type") == "dblclick" else (
            _FILL_EVEN if step_no % 2 == 0 else _FILL_ODD
        )

        data = [step_no, ts_str, "", desc, coord_str, ""]
        for col, val in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border    = _BORDER
            cell.fill      = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col == 1:
                cell.font      = XFont(name="メイリオ", bold=True, size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col in (2, 5):
                cell.alignment = Alignment(horizontal="center", vertical="top")

        # Row height to fit thumbnail
        ws.row_dimensions[row].height = _ROW_HEIGHT

        # Embed annotated thumbnail at column C
        if thumb_png:
            img        = XLImage(io.BytesIO(thumb_png))
            img.width  = _THUMB_W
            img.height = _THUMB_H
            ws.add_image(img, f"{get_column_letter(3)}{row}")

    def _write_log_sheet(self, ws, events, fps):
        headers = ["#", "時刻 (秒)", "MM:SS", "X", "Y", "種別", "UI変化px"]
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.font   = XFont(bold=True, color="FFFFFF")
            c.fill   = _FILL_HEAD
            c.border = _BORDER

        for i, ev in enumerate(events, start=1):
            mm, ss = divmod(int(ev["time_sec"]), 60)
            kind   = "ダブルクリック" if ev.get("type") == "dblclick" else "クリック"
            row    = [i, ev["time_sec"], f"{mm:02d}:{ss:02d}",
                      ev["x"], ev["y"], kind, ev.get("diff_pixels", "")]
            for col, val in enumerate(row, start=1):
                c = ws.cell(row=i + 1, column=col, value=val)
                c.border = _BORDER

        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 14

    # ── Image annotation ─────────────────────────────────────────────────────

    def _make_thumbnail(self, frame: np.ndarray, x: int, y: int) -> bytes:
        """
        Returns PNG bytes of an annotated thumbnail:
          • 赤枠 (red rectangle) around click area
          • Crosshair at exact click point
          • Small "クリック" label above the box
        """
        h_orig, w_orig = frame.shape[:2]

        # Red box boundaries (in original resolution)
        bx1 = max(0,        x - _BOX_MARGIN)
        by1 = max(0,        y - _BOX_MARGIN)
        bx2 = min(w_orig,   x + _BOX_MARGIN)
        by2 = min(h_orig,   y + _BOX_MARGIN)

        # Convert BGR → RGB for PIL
        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        pil = Image.fromarray(rgb)
        draw = ImageDraw.Draw(pil)

        # 赤枠 — draw 3-pixel-wide rectangle
        RED = (210, 30, 30)
        for t in range(3):
            draw.rectangle(
                [bx1 - t, by1 - t, bx2 + t, by2 + t],
                outline=RED,
            )

        # Crosshair at exact click
        CROSS = 14
        draw.line([(x - CROSS, y), (x + CROSS, y)], fill=RED, width=2)
        draw.line([(x, y - CROSS), (x, y + CROSS)], fill=RED, width=2)

        # "クリック" label — small box above the red border
        label = "クリック"
        lx = max(0, bx1)
        ly = max(0, by1 - 22)
        # Red background pill
        draw.rectangle([lx, ly, lx + 72, ly + 18], fill=RED)
        # White text (fallback: default PIL font if CJK font not found)
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 13)
        except Exception:
            font = ImageFont.load_default()
        draw.text((lx + 4, ly + 2), label, fill=(255, 255, 255), font=font)

        # Scale to thumbnail size
        thumb = pil.resize((_THUMB_W, _THUMB_H), Image.LANCZOS)
        buf   = io.BytesIO()
        thumb.save(buf, format="PNG", optimize=True)
        return buf.getvalue()
