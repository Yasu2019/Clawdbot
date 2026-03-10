"""
Report Generator — PDF and Excel report output with optional local LLM comments.
"""
import os
import json
import requests
from pathlib import Path
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment  # noqa: F401 (PatternFill used inline)


_OLLAMA_URL = os.getenv("OLLAMA_BASE_URL", "http://ollama:11434")
_LLM_MODEL = "qwen2.5-coder:7b"  # fast local model; falls back gracefully


def _ask_ollama(prompt: str, max_tokens: int = 400) -> str:
    """Call local Ollama for LLM comment. Returns empty string on any error."""
    try:
        resp = requests.post(
            f"{_OLLAMA_URL}/api/generate",
            json={"model": _LLM_MODEL, "prompt": prompt, "stream": False,
                  "options": {"temperature": 0.3, "num_predict": max_tokens}},
            timeout=60,
        )
        resp.raise_for_status()
        return resp.json().get("response", "").strip()
    except Exception:
        return ""


class ReportGenerator:
    def __init__(self, template: dict):
        self.template = template

    def generate_pdf(self, metrics: dict, labels: list[dict], project_dir: Path) -> Path:
        """Generate PDF report with KPIs, timeline, waste patterns, and ergo info."""
        pdf_path = project_dir / "report.pdf"
        doc = SimpleDocTemplate(str(pdf_path), pagesize=A4,
                                leftMargin=15*mm, rightMargin=15*mm,
                                topMargin=15*mm, bottomMargin=15*mm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("title_custom", parent=styles["Heading1"], fontSize=16)
        h2_style = ParagraphStyle("h2_custom", parent=styles["Heading2"], fontSize=12)
        body_style = styles["BodyText"]

        elements = []

        # Title
        elements.append(Paragraph("WorkStudy AI Analysis Report", title_style))
        elements.append(Paragraph(f"Template: {self.template.get('label', 'N/A')}", body_style))
        elements.append(Spacer(1, 10*mm))

        # KPI Table
        elements.append(Paragraph("KPI Summary", h2_style))
        kpi = metrics.get("kpi", {})
        focus = self.template.get("focus_kpi", [])
        kpi_data = [["KPI", "Value"]]
        for k in focus:
            val = kpi.get(k, "N/A")
            if isinstance(val, float):
                val = f"{val:.3f}"
            kpi_data.append([k, str(val)])

        t = Table(kpi_data, colWidths=[70*mm, 50*mm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 8*mm))

        # Confidence Summary
        avg_conf  = kpi.get("avg_confidence", 0.0)
        low_ratio = kpi.get("low_confidence_ratio", 0.0)
        hi_ratio  = kpi.get("high_confidence_ratio", 0.0)
        if avg_conf > 0:
            elements.append(Paragraph("判定信頼度サマリ", h2_style))
            conf_rows = [
                ["指標", "値", "判定"],
                ["平均信頼度",       f"{avg_conf:.1%}",
                 "良好" if avg_conf >= 0.75 else ("要確認" if avg_conf >= 0.45 else "低信頼")],
                ["高信頼セグメント (≥75%)", f"{hi_ratio:.1%}",
                 "✓" if hi_ratio >= 0.6 else "△"],
                ["低信頼セグメント (<45%)", f"{low_ratio:.1%}",
                 "要再撮影" if low_ratio > 0.3 else "許容範囲"],
            ]
            tc = Table(conf_rows, colWidths=[55*mm, 30*mm, 35*mm])
            tc.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ]))
            elements.append(tc)
            elements.append(Spacer(1, 6*mm))

        # LLM Commentary (local Ollama — zero API cost, skipped if unavailable)
        waste_summary = "; ".join(w["description"] for w in metrics.get("waste_fired", [])[:3]) or "none"
        ergo_trunk = metrics.get("ergo", {}).get("trunk_risk_ratio", 0)
        ergo_shoulder = metrics.get("ergo", {}).get("shoulder_risk_ratio", 0)
        llm_prompt = (
            f"You are a manufacturing process improvement expert.\n"
            f"Process: {self.template.get('label', 'unknown')}\n"
            f"Key KPIs: {json.dumps({k: kpi.get(k) for k in focus[:4]}, ensure_ascii=False)}\n"
            f"Waste patterns triggered: {waste_summary}\n"
            f"Ergonomic risk — trunk: {ergo_trunk:.1%}, shoulder: {ergo_shoulder:.1%}\n"
            f"Write 3 concise bullet-point improvement suggestions in Japanese. Be specific and actionable."
        )
        llm_comment = _ask_ollama(llm_prompt)
        if llm_comment:
            elements.append(Paragraph("AI改善提案 (ローカルLLM)", h2_style))
            for line in llm_comment.split("\n"):
                line = line.strip()
                if line:
                    elements.append(Paragraph(line, body_style))
            elements.append(Spacer(1, 5*mm))

        # Waste Patterns
        waste = metrics.get("waste_fired", [])
        if waste:
            elements.append(Paragraph("Waste Patterns Triggered", h2_style))
            for w in waste:
                elements.append(Paragraph(
                    f"  {w['description']} ({w['metric']}={w['actual_value']} > {w['threshold']})",
                    body_style))
                elements.append(Paragraph(f"    Suggestion: {w['suggestion']}", body_style))
            elements.append(Spacer(1, 5*mm))

        # Ergo
        ergo = metrics.get("ergo", {})
        elements.append(Paragraph("Ergonomic Assessment", h2_style))
        elements.append(Paragraph(
            f"  Trunk risk ratio: {ergo.get('trunk_risk_ratio', 0):.1%} "
            f"(threshold: {ergo.get('trunk_threshold_deg', 40)} deg)", body_style))
        elements.append(Paragraph(
            f"  Shoulder risk ratio: {ergo.get('shoulder_risk_ratio', 0):.1%} "
            f"(threshold: {ergo.get('shoulder_threshold_deg', 60)} deg)", body_style))
        elements.append(Spacer(1, 8*mm))

        # MOST Sequence Table
        most = metrics.get("most", {})
        most_rows = most.get("summary_rows", [])
        if most_rows:
            elements.append(Spacer(1, 8*mm))
            elements.append(Paragraph("BasicMOST シーケンス分析", h2_style))

            # Summary line
            eff_pct = f"{most.get('efficiency', 0):.1%}"
            elements.append(Paragraph(
                f"総TMU: {most.get('total_tmu', 0):.0f}  |  "
                f"NVA-TMU: {most.get('nva_tmu', 0):.0f}  |  "
                f"効率: {eff_pct}  |  "
                f"シーケンス数: {most.get('avg_seq_tmu', 0):.0f} TMU/seq 平均",
                body_style,
            ))
            elements.append(Spacer(1, 3*mm))

            # Index reference note
            elements.append(Paragraph(
                "※ BasicMOST General Move: (A+B+G+A+B+P+A)×10 TMU  "
                "| 1 TMU = 0.036 s  "
                "| A=距離 B=身体 G=つかみ P=位置決め",
                ParagraphStyle("note", parent=body_style, fontSize=7, textColor=colors.grey),
            ))
            elements.append(Spacer(1, 2*mm))

            most_hdr = ["#", "タイプ", "A", "B", "G", "P", "TMU", "時間(s)"]
            most_data = [most_hdr] + [
                [str(r[0]), r[1], r[2], r[3], r[4], r[5], f"{r[6]:.0f}", f"{r[7]:.1f}"]
                for r in most_rows[:40]
            ]
            tm = Table(most_data, colWidths=[10*mm, 32*mm, 12*mm, 12*mm, 12*mm, 12*mm, 16*mm, 16*mm])
            tm.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (2, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ]))
            elements.append(tm)
            elements.append(Spacer(1, 8*mm))

        # Timeline
        elements.append(Paragraph("Motion Timeline (Therblig)", h2_style))
        tl_data = [["#", "Start(s)", "End(s)", "Therblig", "Dur(s)", "TMU", "Conf", "NVA"]]
        tl_styles = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ]
        for row_i, lbl in enumerate(labels[:30], start=1):
            conf = lbl.get("confidence", 0.0)
            tl_data.append([
                str(lbl["segment_id"]),
                f"{lbl['start_sec']:.1f}",
                f"{lbl['end_sec']:.1f}",
                lbl.get("label_jp", lbl["label"]),
                f"{lbl['duration_sec']:.1f}",
                f"{lbl.get('most_tmu', 0):.0f}",
                f"{conf:.0%}",
                "●" if lbl["is_nva"] else "",
            ])
            # Colour-code confidence cell (column 6, 0-indexed)
            if conf >= 0.75:
                tl_styles.append(("BACKGROUND", (6, row_i), (6, row_i), colors.HexColor("#D1FAE5")))
            elif conf >= 0.45:
                tl_styles.append(("BACKGROUND", (6, row_i), (6, row_i), colors.HexColor("#FEF9C3")))
            else:
                tl_styles.append(("BACKGROUND", (6, row_i), (6, row_i), colors.HexColor("#FFEDD5")))
        t2 = Table(tl_data, colWidths=[10*mm, 18*mm, 18*mm, 34*mm, 16*mm, 14*mm, 14*mm, 8*mm])
        t2.setStyle(TableStyle(tl_styles))
        elements.append(t2)

        doc.build(elements)
        return pdf_path

    def generate_xlsx(self, metrics: dict, labels: list[dict], project_dir: Path) -> Path:
        """Generate Excel report with KPI and timeline sheets."""
        xlsx_path = project_dir / "report.xlsx"
        wb = openpyxl.Workbook()

        # KPI Sheet
        ws_kpi = wb.active
        ws_kpi.title = "KPI"
        header_fill = PatternFill("solid", fgColor="2563EB")
        header_font = Font(color="FFFFFF", bold=True)

        ws_kpi.append(["KPI", "Value"])
        for cell in ws_kpi[1]:
            cell.fill = header_fill
            cell.font = header_font

        kpi = metrics.get("kpi", {})
        for k in self.template.get("focus_kpi", []):
            val = kpi.get(k, "N/A")
            ws_kpi.append([k, val])

        ws_kpi.column_dimensions["A"].width = 30
        ws_kpi.column_dimensions["B"].width = 20

        # Timeline Sheet
        ws_tl = wb.create_sheet("Timeline")
        ws_tl.append(["#", "Start(s)", "End(s)", "Label", "Label(JP)", "Duration(s)",
                       "A", "B", "G", "P", "TMU", "Confidence", "vis_ratio", "NVA"])
        for cell in ws_tl[1]:
            cell.fill = header_fill
            cell.font = header_font

        green_fill  = PatternFill("solid", fgColor="D1FAE5")
        yellow_fill = PatternFill("solid", fgColor="FEF9C3")
        orange_fill = PatternFill("solid", fgColor="FFEDD5")

        for row_i, lbl in enumerate(labels, start=2):
            conf = lbl.get("confidence", 0.0)
            ws_tl.append([
                lbl["segment_id"], lbl["start_sec"], lbl["end_sec"],
                lbl["label"], lbl.get("label_jp", ""), lbl["duration_sec"],
                lbl.get("most_A", ""), lbl.get("most_B", ""),
                lbl.get("most_G", ""), lbl.get("most_P", ""),
                lbl.get("most_tmu", ""),
                f"{conf:.1%}",
                f"{lbl.get('vis_ratio', 0):.1%}",
                "Yes" if lbl["is_nva"] else "",
            ])
            # Colour confidence cell (column 12 = L)
            conf_cell = ws_tl.cell(row=row_i, column=12)
            if conf >= 0.75:
                conf_cell.fill = green_fill
            elif conf >= 0.45:
                conf_cell.fill = yellow_fill
            else:
                conf_cell.fill = orange_fill

        # MOST Sequences Sheet
        most = metrics.get("most", {})
        if most.get("sequences"):
            ws_most = wb.create_sheet("MOSTシーケンス")
            ws_most.append(["#", "タイプ", "A", "B", "G", "P", "TMU", "時間(s)", "NVA"])
            for cell in ws_most[1]:
                cell.fill = header_fill
                cell.font = header_font
            for row in most.get("summary_rows", []):
                ws_most.append(row)
            # Summary row
            ws_most.append([])
            ws_most.append(["合計", "",
                            "", "", "", "",
                            most.get("total_tmu", ""),
                            round(most.get("total_tmu", 0) * 0.036, 1),
                            ""])
            ws_most.append(["NVA TMU", "", "", "", "", "",
                            most.get("nva_tmu", ""), "", ""])
            ws_most.append(["効率", "", "", "", "", "",
                            f"{most.get('efficiency', 0):.1%}", "", ""])
            for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
                ws_most.column_dimensions[col].width = 14

        # Waste Sheet
        ws_waste = wb.create_sheet("Waste Patterns")
        ws_waste.append(["ID", "Description", "Metric", "Actual", "Threshold", "Suggestion"])
        for cell in ws_waste[1]:
            cell.fill = header_fill
            cell.font = header_font

        for w in metrics.get("waste_fired", []):
            ws_waste.append([w["id"], w["description"], w["metric"],
                             w["actual_value"], w["threshold"], w["suggestion"]])

        wb.save(str(xlsx_path))
        return xlsx_path
