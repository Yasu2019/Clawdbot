"""
FactoryProcedureWriter — 工場作業動画から作業標準書Excelを生成する。

各セグメント（動作ラベル）の代表フレームをキャプチャし、
MediaPipeの骨格ランドマークをOpenCVでオーバーレイして埋め込む。

出力シート:
  「作業標準書」 : Step / 時刻 / スクリーンショット(骨格付き) / 動作ラベル / 時間 / NVA / 備考
  「KPIサマリ」  : テンプレートのfocus_kpi一覧
  「タイムライン」: 全セグメントのリスト
"""

import io
import cv2
import numpy as np
from pathlib import Path
from datetime import datetime

from PIL import Image, ImageDraw, ImageFont

import openpyxl
from openpyxl.styles import Font as XFont, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


# ── Excel styling ─────────────────────────────────────────────────────────────
_FILL_TITLE  = PatternFill("solid", fgColor="1E3A5F")
_FILL_HEAD   = PatternFill("solid", fgColor="2563EB")
_FILL_NVA    = PatternFill("solid", fgColor="FEE2E2")   # red tint for NVA steps
_FILL_EVEN   = PatternFill("solid", fgColor="F0F7FF")
_FILL_ODD    = PatternFill("solid", fgColor="FFFFFF")
_SIDE        = Side(style="thin", color="CCCCCC")
_BORDER      = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)

_THUMB_W     = 320
_THUMB_H     = 200
_ROW_HEIGHT  = _THUMB_H * 0.75 + 8   # Excel points

# ── MediaPipe pose connections (33 landmarks, standard skeleton) ──────────────
_POSE_CONNECTIONS = [
    # Face
    (0, 1), (1, 2), (2, 3), (3, 7),
    (0, 4), (4, 5), (5, 6), (6, 8),
    (9, 10),
    # Upper body
    (11, 12),
    (11, 13), (13, 15), (15, 17), (17, 19), (19, 15), (15, 21),
    (12, 14), (14, 16), (16, 18), (18, 20), (20, 16), (16, 22),
    (11, 23), (12, 24), (23, 24),
    # Lower body
    (23, 25), (25, 27), (27, 29), (29, 31), (31, 27),
    (24, 26), (26, 28), (28, 30), (30, 32), (32, 28),
]

# Landmark groups for colour coding
_HAND_IDX     = {15, 16, 17, 18, 19, 20, 21, 22}
_SHOULDER_IDX = {11, 12}
_HIP_IDX      = {23, 24}


class FactoryProcedureWriter:
    """工場動作セグメントのキャプチャ + 骨格描画 → 作業標準書Excel生成。"""

    # ── Public API ────────────────────────────────────────────────────────────

    def write_excel(
        self,
        labels: list[dict],
        pose_data: list[dict],
        video_path: str,
        project_dir: Path,
        template: dict,
        title: str = "作業標準書",
        progress_cb=None,
    ) -> Path:
        """
        Generate work-instruction Excel for factory analysis.

        Args:
            labels:       Output of TherbligLabeler.label().
            pose_data:    Output of PoseEstimator.process() (per-frame landmarks).
            video_path:   Original video file.
            project_dir:  Output directory.
            template:     Template dict (for focus_kpi, waste_patterns etc.).
            title:        Document title.
            progress_cb:  Optional (frac, desc) callback for real-time updates.

        Returns:
            Path to procedure.xlsx
        """
        xlsx_path = project_dir / "procedure.xlsx"
        wb        = openpyxl.Workbook()

        # ── Sheet 1: 作業標準書 ───────────────────────────────────────────────
        ws = wb.active
        ws.title = "作業標準書"
        self._write_title(ws, title)
        self._write_col_headers(ws)
        self._set_col_widths(ws)

        cap   = cv2.VideoCapture(video_path)
        fps   = cap.get(cv2.CAP_PROP_FPS) or 30.0
        total = max(len(labels), 1)

        data_row = 3
        for step_no, lbl in enumerate(labels, start=1):
            if progress_cb:
                progress_cb(
                    step_no / total,
                    f"手順書生成中... Step {step_no}/{total}  ({lbl.get('label', '')})",
                )

            # Representative frame: middle of segment
            rep_frame = (lbl["start_frame"] + lbl["end_frame"]) // 2
            cap.set(cv2.CAP_PROP_POS_FRAMES, rep_frame)
            ret, frame = cap.read()
            if not ret:
                continue

            # Skeleton landmarks for this frame
            lm_list = (
                pose_data[rep_frame]["landmarks"]
                if rep_frame < len(pose_data)
                else []
            )

            thumb_png = self._make_thumbnail(frame, lm_list, lbl)
            self._write_step_row(ws, step_no, lbl, fps, thumb_png, data_row)
            data_row += 1

        cap.release()

        # ── Sheet 2: KPIサマリ ────────────────────────────────────────────────
        ws_kpi = wb.create_sheet("KPIサマリ")
        self._write_kpi_sheet(ws_kpi, labels, template)

        # ── Sheet 3: タイムライン ─────────────────────────────────────────────
        ws_tl = wb.create_sheet("タイムライン")
        self._write_timeline_sheet(ws_tl, labels)

        wb.save(str(xlsx_path))
        return xlsx_path

    # ── Excel writing helpers ─────────────────────────────────────────────────

    def _write_title(self, ws, title: str):
        ws.merge_cells("A1:H1")
        c = ws["A1"]
        c.value     = f"{title}  —  生成日時: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        c.font      = XFont(name="Meiryo", size=14, bold=True, color="FFFFFF")
        c.fill      = _FILL_TITLE
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

    def _write_col_headers(self, ws):
        headers = ["Step", "時刻", "スクリーンショット (骨格付き)", "動作ラベル",
                   "持続時間(s)", "信頼度", "NVA", "備考・改善提案"]
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=2, column=col, value=h)
            c.font      = XFont(name="Meiryo", size=10, bold=True, color="FFFFFF")
            c.fill      = _FILL_HEAD
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = _BORDER
        ws.row_dimensions[2].height = 24

    def _set_col_widths(self, ws):
        widths = {"A": 7, "B": 10, "C": 46, "D": 22, "E": 12, "F": 10, "G": 8, "H": 30}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

    def _write_step_row(self, ws, step_no, lbl, fps, thumb_png, row):
        mm, ss   = divmod(int(lbl["start_sec"]), 60)
        ts_str   = f"{mm:02d}:{ss:02d}"
        label_jp = lbl.get("label_jp", lbl.get("label", ""))
        dur      = round(lbl.get("duration_sec", 0), 1)
        is_nva   = lbl.get("is_nva", False)
        conf     = lbl.get("confidence", 0.0)
        conf_str = f"{conf:.0%}"
        fill     = _FILL_NVA if is_nva else (_FILL_EVEN if step_no % 2 == 0 else _FILL_ODD)

        # Confidence cell fill: green/yellow/orange
        if conf >= 0.75:
            conf_fill = PatternFill("solid", fgColor="D1FAE5")   # green tint
        elif conf >= 0.45:
            conf_fill = PatternFill("solid", fgColor="FEF9C3")   # yellow tint
        else:
            conf_fill = PatternFill("solid", fgColor="FFEDD5")   # orange tint

        row_data = [step_no, ts_str, "", label_jp, dur, conf_str, "●" if is_nva else "", ""]
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.border    = _BORDER
            c.fill      = conf_fill if col == 6 else fill
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if col == 1:
                c.font      = XFont(bold=True, size=12)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col in (2, 5, 6):
                c.alignment = Alignment(horizontal="center", vertical="center")
                if col == 6:
                    if conf < 0.45:
                        c.font = XFont(color="C05000", bold=True)
                    elif conf >= 0.75:
                        c.font = XFont(color="166534", bold=True)
            elif col == 7 and is_nva:
                c.font = XFont(color="CC0000", bold=True)

        ws.row_dimensions[row].height = _ROW_HEIGHT

        if thumb_png:
            try:
                img_buf    = io.BytesIO(thumb_png)
                img        = XLImage(img_buf)
                img.width  = _THUMB_W
                img.height = _THUMB_H
                ws.add_image(img, f"{get_column_letter(3)}{row}")
            except Exception:
                pass  # skip broken thumbnail; rest of file remains valid

    def _write_kpi_sheet(self, ws, labels, template):
        ws.merge_cells("A1:C1")
        c = ws["A1"]
        c.value = "KPIサマリ"
        c.font  = XFont(bold=True, size=13, color="FFFFFF")
        c.fill  = _FILL_HEAD
        ws.row_dimensions[1].height = 24

        total_sec = sum(l.get("duration_sec", 0) for l in labels)
        nva_sec   = sum(l.get("duration_sec", 0) for l in labels if l.get("is_nva"))
        va_sec    = total_sec - nva_sec

        rows = [
            ("総セグメント数",    len(labels)),
            ("総動作時間 (秒)",   round(total_sec, 1)),
            ("付加価値時間 (秒)", round(va_sec, 1)),
            ("非付加価値時間(秒)", round(nva_sec, 1)),
            ("NVA比率",          f"{nva_sec/total_sec:.1%}" if total_sec > 0 else "N/A"),
        ]
        for i, (k, v) in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=k).border = _BORDER
            ws.cell(row=i, column=2, value=v).border = _BORDER

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 16

    def _write_timeline_sheet(self, ws, labels):
        headers = ["#", "開始(s)", "終了(s)", "時刻", "ラベル", "ラベル(JP)", "時間(s)", "NVA"]
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.font   = XFont(bold=True, color="FFFFFF")
            c.fill   = _FILL_HEAD
            c.border = _BORDER

        for i, lbl in enumerate(labels, start=1):
            mm, ss = divmod(int(lbl["start_sec"]), 60)
            row = [
                i,
                round(lbl["start_sec"], 2),
                round(lbl["end_sec"], 2),
                f"{mm:02d}:{ss:02d}",
                lbl.get("label", ""),
                lbl.get("label_jp", ""),
                round(lbl.get("duration_sec", 0), 2),
                "●" if lbl.get("is_nva") else "",
            ]
            for col, val in enumerate(row, start=1):
                c = ws.cell(row=i + 1, column=col, value=val)
                c.border = _BORDER
                if lbl.get("is_nva"):
                    c.fill = _FILL_NVA

        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 14

    # ── Frame annotation ──────────────────────────────────────────────────────

    def _make_thumbnail(
        self,
        frame: np.ndarray,
        landmarks: list[dict],
        lbl: dict,
    ) -> bytes:
        """
        Returns PNG bytes of annotated frame:
          - Skeleton overlay (connections + colour-coded joints)
          - Motion label banner at bottom
        """
        h_orig, w_orig = frame.shape[:2]

        # Draw skeleton directly on BGR frame
        annotated = self._draw_skeleton(frame.copy(), landmarks, w_orig, h_orig)

        # Convert to PIL for the label banner
        rgb  = cv2.cvtColor(annotated, cv2.COLOR_BGR2RGB)
        pil  = Image.fromarray(rgb)
        draw = ImageDraw.Draw(pil)

        # Bottom banner: semi-transparent dark strip with motion label
        banner_h = 36
        banner   = Image.new("RGBA", (w_orig, banner_h), (0, 0, 0, 180))
        pil_rgba = pil.convert("RGBA")
        pil_rgba.paste(banner, (0, h_orig - banner_h), banner)
        pil      = pil_rgba.convert("RGB")
        draw     = ImageDraw.Draw(pil)

        label_text = lbl.get("label_jp") or lbl.get("label", "")
        nva_flag   = "  [NVA]" if lbl.get("is_nva") else ""
        dur_text   = f"  {lbl.get('duration_sec', 0):.1f}s"
        banner_str = f"  {label_text}{nva_flag}{dur_text}"

        try:
            font = ImageFont.truetype(
                "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 18
            )
        except Exception:
            font = ImageFont.load_default()

        label_color = (255, 80, 80) if lbl.get("is_nva") else (100, 255, 100)
        draw.text((0, h_orig - banner_h + 6), banner_str, fill=label_color, font=font)

        # Scale to thumbnail
        thumb = pil.resize((_THUMB_W, _THUMB_H), Image.LANCZOS)
        buf   = io.BytesIO()
        thumb.save(buf, format="PNG", optimize=True)
        return buf.getvalue()

    @staticmethod
    def _draw_skeleton(
        frame: np.ndarray,
        landmarks: list[dict],
        w: int,
        h: int,
    ) -> np.ndarray:
        """Draw MediaPipe pose skeleton onto a BGR frame."""
        if not landmarks:
            return frame

        # Draw connections
        for a_i, b_i in _POSE_CONNECTIONS:
            if a_i >= len(landmarks) or b_i >= len(landmarks):
                continue
            a, b = landmarks[a_i], landmarks[b_i]
            if a["visibility"] < 0.35 or b["visibility"] < 0.35:
                continue
            ax, ay = int(a["x"] * w), int(a["y"] * h)
            bx, by = int(b["x"] * w), int(b["y"] * h)
            cv2.line(frame, (ax, ay), (bx, by), (180, 180, 180), 2, cv2.LINE_AA)

        # Draw joints (colour-coded by body part)
        for i, lm in enumerate(landmarks):
            if lm["visibility"] < 0.35:
                continue
            x, y = int(lm["x"] * w), int(lm["y"] * h)
            if i in _HAND_IDX:
                color, r = (0, 220, 0), 7      # green — hands (most important for work study)
            elif i in _SHOULDER_IDX:
                color, r = (0, 160, 255), 6    # orange — shoulders
            elif i in _HIP_IDX:
                color, r = (255, 200, 0), 5    # yellow — hips
            else:
                color, r = (220, 220, 220), 4  # white — other

            cv2.circle(frame, (x, y), r, color, -1, cv2.LINE_AA)
            cv2.circle(frame, (x, y), r + 1, (0, 0, 0), 1, cv2.LINE_AA)  # outline

        return frame
